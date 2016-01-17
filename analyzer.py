import sqlite3
import operator
from collections import defaultdict
from openpyxl import Workbook

conn = sqlite3.connect('firstquarter.db')   # CHANGE DEPENDING ON DATABASE
c = conn.cursor()

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = "First Quarter"                  # CHANGE DEPENDING ON DATABASE

# create titles
ws['A1'].value = "Username"
ws['B1'].value = "Posts"
ws['C1'].value = "Flair 1"
ws['D1'].value = "Flair 2"
ws['E1'].value = "Word"
ws['F1'].value = "Word Count"

prev_user = ""
post_count = 0
user_total = 0
delete_num = 0

# unique users (EXCLUDING [deleted]) & post count (INCLUDING [deleted])
for row in c.execute('SELECT username FROM comments ORDER BY username'):
    if row[0] != prev_user and "/u/None" not in row[0]:
        prev_user = row[0]
        user_total += 1
    if "/u/None" in row[0]:
        delete_num += 1
    post_count += 1
print("There were " + str(user_total) + " individuals users contributing " + str(post_count) + " total comments.")
print("Of the " + str(post_count) + " total posts, " + str(delete_num) + " have been deleted.\n")

rown = 2 # row number iterator

# Users and their post amount
for row in c.execute('SELECT username, count(*) FROM comments GROUP BY username ORDER BY username'):
    ws.cell(row=rown, column=1).value = row[0]
    ws.cell(row=rown, column=2).value = int(float(row[1]))
    rown += 1

# FLAIRS
bama_fan = 0
clem_fan = 0
bamaclem = 0
sec_fan  = 0
acc_fan  = 0

bama_flair = ['Alabama Band', 'Crimson Tide']
sec_flair  = ['Crimson Tide', 'Kentucky Wildcats', 'Carolina Gamecocks', 'Arkansas Razorbacks'
                'LSU', 'Volunteer', 'Ole Miss', 'Auburn', 'Mississippi State', 'Texas A&M'
                'Florida Gators', 'Missouri Tigers', 'Vanderbilt', 'Georgia Bulldog', 'SEC']
acc_flair = ['Boston College', 'Georgia Tech', 'Carolina State Wolf', 'Virginia Tech', 'Clemson',
                'Louisville Card', 'Pittsburg Panthers', 'Wake Forest Demon', 'Duke Blue Devils',
                'Miami Hurricanes', 'Syracuse', 'Florida State', 'North Carlonia Tar', 'Virginia Caveliers', 'ACC']

rown = 2
for row in c.execute ('SELECT flair1, flair2 FROM comments GROUP BY username ORDER BY username'):
    
    flair1 = row[0]
    flair2 = row[1]
    
    ws.cell(row=rown, column=3).value = flair1
    ws.cell(row=rown, column=4).value = flair2
    rown = rown + 1
    
    # Check Bama Fan
    if any(bama in flair1 for bama in bama_flair) or any(bama in flair2 for bama in bama_flair):
        bama_fan = bama_fan + 1
    # Check Clemson Fan
    if 'Clemson' in flair1 or 'Clemson' in flair2:
        clem_fan = clem_fan + 1
    # Check if Bama / Clemson or Clemson / Bama
    if ('Clemson' in flair1 and any(bama in flair2 for bama in bama_flair)) or any(bama in flair1 for bama in bama_flair) and 'Clemson' in flair2:
        bama_fan = bama_fan - 1
        clem_fan = clem_fan - 1
        bamaclem = bamaclem + 1

    # Check SEC flairs
    if any(sec in flair1 for sec in sec_flair) or any(sec in flair2 for sec in sec_flair):
        sec_fan = sec_fan + 1
    # Check ACC flairs
    if any(acc in flair1 for acc in acc_flair) or any(acc in flair2 for acc in acc_flair):
        acc_fan = acc_fan + 1

# Comment bodies and individual word frequency
dict = defaultdict(int)
bama_w = 0
clem_w = 0
fuck_w = 0
shit_w = 0
henry_w = 0
watson_w = 0
touchdown_w = 0
#onside_w = 0    # uncomment for 4th quarter

for row in c.execute ('SELECT comment FROM comments'):
    if "bama" in row[0].lower(): # all to lower case, for ease in checking
        bama_w += 1
    if "clemson" in row[0].lower() or "clempson" in row[0].lower():
        clem_w += 1
    if "fuck" in row[0].lower():
        fuck_w += 1
    if "shit" in row[0].lower():
        shit_w += 1
    if "derrick" in row[0].lower() or "henry" in row[0].lower():
        henry_w += 1
    if "deshaun" in row[0].lower() or "watson" in row[0].lower():
        watson_w += 1
    if "touchdown" in row[0].lower():
        touchdown_w += 1
#    if "onside" in row[0].lower(): # uncomment for fourth quarter
#        onside_w += 1
    row_s = row[0].split()
    for word in row_s:
        word = word.lower()
        dict[word] += 1

print("Comments related to Alabama: " + str(bama_w))
print("Comments related to Clemson: " + str(clem_w))
print("Comments saying a variant of \"fuck\": " + str(fuck_w))
print("Comments saying a variant of \"shit\": " + str(shit_w))
print("Comments mentioning Derrick Henry: " + str(henry_w))
print("Comments mentioning Deshaun Watson: " + str(watson_w))
print("Comments talking about a touchdown: " + str(touchdown_w))
#print("Comments about an onside kick: " + str(onside_w))
sorted = sorted(dict.items(), key=operator.itemgetter(1))
rown = 2

for item in sorted:
    ws.cell(row=rown, column=5).value = item[0]
    ws.cell(row=rown, column=6).value = item[1]
    rown += 1

wb.save('first_results.xlsx')   # CHANGE DEPENDING ON DATABASE
