import sqlite3
import operator
import datetime, calendar
from collections import defaultdict
from openpyxl import Workbook, load_workbook

# Created by Maxwell Smith - Running Python 2.7.10
# This file holds the function definitions for every function used in the analyzer.py file.
# This allows for more readability in the analyzer.py file as well as clarity as to what functions
# are doing what.

# connect to sqlite server
conn = sqlite3.connect('cfb.db')
c = conn.cursor()

# Sets up the Excel workbook sheet
# Returns nothing
def setup_workbook(database):
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = str(database)

    # create titles
    ws['A1'].value = "Username"
    ws['B1'].value = "Posts"
    ws['C1'].value = "Flair 1"
    ws['D1'].value = "Flair 2"
    ws['E1'].value = "Word"
    ws['F1'].value = "Word Count"
    ws['G1'] = "Flair Breakdown"
    ws['G2'] = "Bama Flairs"
    ws['G3'] = "Clemson Flairs"
    ws['G4'] = "Michigan St Flairs"
    ws['G5'] = "Oklahoma Sooners Flairs"
    ws['G6'] = "Bama / Clemson Flairs"
    ws['G7'] = "SEC Flairs"
    ws['G8'] = "ACC Flairs"
    ws['G9'] = "Big 12 Flairs"
    ws['G10'] = "Big 10 Flairs"
    ws['G11'] = "Pac 12 Flairs"
    ws['H1'] = "Flair Count"
    ws['I1'] = "Time of Comment"
    ws['J1'] = "Posts per Minute"
    
    # Save
    wb.save(database + '.xlsx')

# Analyzes the unique users and unique user post count. Places them into spreadsheet
# Returns nothing
def unique_users(database):
    # Reopen Workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    prev_user  = ""
    user_total = 0
    post_count = 0
    delete_num = 0
    execute    = 'SELECT username FROM ' + database + ' ORDER BY username'

    for row in c.execute(execute):
        if row[0] != prev_user and "/u/None" not in row[0]:
            prev_user = row[0]
            user_total += 1
        if "/u/None" in row[0]:
            delete_num += 1
        post_count += 1
    
    print("There were " + str(user_total) + " individuals users contributing " + str(post_count) + " total comments.")
    print("Of the " + str(post_count) + " total posts, " + str(delete_num) + " have been deleted.\n")

    rown = 2 # row number iterator
    # Users and their post amount
    for row in c.execute('SELECT username, count(*) FROM ' + database + ' GROUP BY username ORDER BY username'):
        ws.cell(row=rown, column=1).value = row[0]
        ws.cell(row=rown, column=2).value = int(float(row[1]))
        rown += 1

    # Resave workbook
    wb.save(database + '.xlsx')

def flairs(database):
    # Reopen workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    bama_fan   = 0
    clem_fan   = 0
    bamaclem   = 0
    sec_fan    = 0
    acc_fan    = 0
    michst_fan = 0
    sooner_fan = 0
    big12_fan  = 0
    big10_fan  = 0
    pac12_fan  = 0

    bama_flair = ['Alabama Band', 'Crimson Tide']

    sec_flair  = ['Crimson Tide', 'Kentucky Wildcats', 'Carolina Gamecocks', 'Arkansas Razorbacks'
                  'LSU', 'Volunteer', 'Ole Miss', 'Auburn', 'Mississippi State', 'Texas A&M'
                  'Florida Gators', 'Missouri Tigers', 'Vanderbilt', 'Georgia Bulldog', 'SEC']

    acc_flair = ['Boston College', 'Georgia Tech', 'Carolina State Wolf', 'Virginia Tech', 'Clemson',
                 'Louisville Card', 'Pittsburg Panthers', 'Wake Forest Demon', 'Duke Blue Devils',
                 'Miami Hurricanes', 'Syracuse', 'Florida State', 'North Carlonia Tar', 'Virginia Caveliers', 'ACC']

    big12_flair = ['Baylor Bears', 'Oklahoma Sooners', 'Oklahoma Bandwagon' 'Texas Longhorns',
                   'Iowa State Cyclones','Oklahoma State', 'Texas Tech Red', 'Kansas Jay',
                   'TCU Horned', 'West Virginia', 'Kansas State']

    big10_flair = ['Illini', 'Michigan State', 'Ohio State', 'Wisonsin Bad', 'Indiana Hoosiers',
                   'Penn State Nittany', 'Iowa Hawkeyes', 'Nebraska Cornhuskers', 'Purdue Boilermakers',
                   'Maryland Terrapins', 'Northwestern Wildcats', 'Rutgers Scarlet', 'Michigan Wolverines']
    
    pac12_flair = ['Arizona Wildcats', 'Oregon Ducks', 'USC Trojans', 'Arizona State Sun', 'Oregon State',
                   'Utah Utes', 'California Golden', 'Stanford', 'Washington Huskies',
                   'Colorado Buffaloes', 'UCLA Bruins', 'Washington State']

    rown = 2
    execute = 'SELECT flair1, flair2 FROM ' + database + ' GROUP BY username ORDER BY username'
    for row in c.execute (execute):
        
        flair1 = row[0]
        flair2 = row[1]
        
        ws.cell(row=rown, column=3).value = flair1
        ws.cell(row=rown, column=4).value = flair2
        rown = rown + 1
        
        # Check Bama Fan
        if any(bama in flair1 for bama in bama_flair) or any(bama in flair2 for bama in bama_flair):
            bama_fan = bama_fan + 1
        # Check Clemson Fan
        if 'Clemson' in flair1 or 'Clemson' in flair2:
            clem_fan = clem_fan + 1
        # Check if Bama / Clemson or Clemson / Bama
        if ('Clemson' in flair1 and any(bama in flair2 for bama in bama_flair)) or any(bama in flair1 for bama in bama_flair) and 'Clemson' in flair2:
            bama_fan = bama_fan - 1
            clem_fan = clem_fan - 1
            bamaclem = bamaclem + 1
        
        # Check Michigan St
        if 'Michigan State' in flair1 or 'Michigan State' in flair2:
            michst_fan += 1
        
        # Check Oklahoma Sooner fan
        if 'Oklahoma Band' in flair1 or 'Oklahoma Band' in flair2 or 'Sooners' in flair1 or 'Sooners' in flair2:
            sooner_fan += 1

        # Check SEC flairs
        if any(sec in flair1 for sec in sec_flair) or any(sec in flair2 for sec in sec_flair):
            sec_fan = sec_fan + 1
        # Check ACC flairs
        if any(acc in flair1 for acc in acc_flair) or any(acc in flair2 for acc in acc_flair):
            acc_fan = acc_fan + 1
        # Check BIG12 flairs
        if any(big12 in flair1 for big12 in big12_flair) or any(big12 in flair2 for big12 in big12_flair):
            big12_fan = big12_fan + 1
        # Check BIG10 flairs
        if any(big10 in flair1 for big10 in big10_flair) or any(big10 in flair2 for big10 in big10_flair):
            big10_fan = big10_fan + 1
        # Check PAC12 flairs
        if any(pac12 in flair1 for pac12 in pac12_flair) or any(pac12 in flair2 for pac12 in pac12_flair):
            pac12_fan = pac12_fan + 1

    fan_arr = [bama_fan, clem_fan, michst_fan, sooner_fan, bamaclem, sec_fan, acc_fan, big12_fan, big10_fan, pac12_fan]
    
    # Place in ws via loop (insteasd of copy/pasting)
    for fan in range(0, len(fan_arr)):
        ws.cell(row=(fan+2), column=8).value = fan_arr[fan]
    
    # Resave Workbook
    wb.save(database + '.xlsx')

def comments(database):
    # Reopen Workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    # Comment bodies and individual word frequency
    dict = defaultdict(int)
    bama_w = 0
    clem_w = 0
    fuck_w = 0
    shit_w = 0
    henry_w = 0
    watson_w = 0
    touchdown_w = 0
    onside_w = 0

    for row in c.execute ('SELECT comment FROM ' + database):
        comment = row[0].lower()
        if "bama" in comment: # all to lower case, for ease in checking
            bama_w += 1
        if "clemson" in comment or "clempson" in comment:
            clem_w += 1
        if "fuck" in comment:
            fuck_w += 1
        if "shit" in comment:
            shit_w += 1
        if "derrick" in comment or "henry" in comment:
            henry_w += 1
        if "deshaun" in comment or "watson" in comment:
            watson_w += 1
        if "touchdown" in comment:
            touchdown_w += 1
        if "onside" in comment: # uncomment for fourth quarter
            onside_w += 1
        row_s = row[0].split()
        for word in row_s:
            word = word.lower()
            dict[word] += 1

    print("Comments related to Alabama: " + str(bama_w))
    print("Comments related to Clemson: " + str(clem_w))
    print("Comments saying a variant of \"fuck\": " + str(fuck_w))
    print("Comments saying a variant of \"shit\": " + str(shit_w))
    print("Comments mentioning Derrick Henry: " + str(henry_w))
    print("Comments mentioning Deshaun Watson: " + str(watson_w))
    print("Comments talking about a touchdown: " + str(touchdown_w))
    print("Comments about an onside kick: " + str(onside_w))

    # Sort and organize dictionary for placement in Worksheet
    sorted_dict = sorted(dict.items(), key=operator.itemgetter(1))

    rown = 2
    for item in sorted_dict:
        ws.cell(row=rown, column=5).value = item[0]
        ws.cell(row=rown, column=6).value = item[1]
        rown += 1

    # Resave Workbook
    wb.save(database + '.xlsx')

def comments_per_minute(database):
    # Reopen Workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active

    comdict = defaultdict(int)
    for row in c.execute('SELECT time FROM ' + database):
        date_time = datetime.datetime.utcfromtimestamp(int(row[0]))
        fmt_time  = date_time.strftime("%D %H:%M")
        fmt_time = str(fmt_time)
        comdict[fmt_time] += 1

    sorted_dict = sorted(comdict.items(), key=operator.itemgetter(1))
    rown=2
    for item in sorted_dict:
        ws.cell(row=rown, column=9).value = item[0]
        ws.cell(row=rown, column=10).value = item[1]
        rown += 1


    # Save Workbook
    wb.save(database + '.xlsx')